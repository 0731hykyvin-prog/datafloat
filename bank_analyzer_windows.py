#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import tkinter.scrolledtext as scrolledtext
import os
import sys
import json
import glob
from decimal import Decimal
from datetime import datetime, timedelta
import threading
import re
from openpyxl import load_workbook
from collections import defaultdict

class BankTransactionAnalyzer:
    def __init__(self, root):
        self.root = root
        self.root.title("银行交易明细分析工具 v3.0")
        self.root.geometry("1200x900")
        self.root.resizable(True, True)
        
        self.config_file = os.path.join(os.path.expanduser("~"), ".bank_analyzer_config.json")
        self.root.configure(bg='#f0f0f0')
        
        self.current_file = None
        self.df = None
        
        self.merge_in_progress = False
        self.analysis_in_progress = False
        self.complete_in_progress = False
        
        # 去重开关（用户选择）
        self.deduplicate = True
        
        self.load_config()
        self.create_widgets()
    
    def load_config(self):
        default_config = {
            "column_mapping": {
                "交易时间": "交易时间",
                "交易金额": "交易金额",
                "交易卡号": "交易卡号",
                "交易账号": "交易账号",
                "账户开户名称": "账户开户名称",
                "开户人证件号码": "开户人证件号码",
                "交易方户名": "交易方户名",
                "交易方账号": "交易方账号",
                "交易方证件号码": "交易方证件号码",
                "收付标志": "收付标志",
                "交易摘要": "交易摘要",
                "查询反馈结果原因": "查询反馈结果原因",
                "交易流水号": "交易流水号"
            },
            "analysis_params": {
                "night_start": 21,
                "night_end": 5,
                "min_amount": 500
            },
            "output_order": [
                "交易流水号",
                "交易卡号",
                "交易账号",
                "账户开户名称",
                "开户人证件号码",
                "交易时间",
                "交易金额"
            ],
            "complete_config": {
                "transaction_file": "",
                "account_file": ""
            },
            "behavior_analysis": {
                "frequency_threshold": 10,
                "amount_threshold": 50000,
                "time_window_days": 30
            }
        }
        
        if os.path.exists(self.config_file):
            try:
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    saved_config = json.load(f)
                    for key in default_config:
                        if key in saved_config:
                            if isinstance(default_config[key], dict):
                                default_config[key].update(saved_config[key])
                            else:
                                default_config[key] = saved_config[key]
            except:
                pass
        
        self.config = default_config
        self.save_config()
    
    def save_config(self):
        try:
            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(self.config, f, ensure_ascii=False, indent=2)
        except:
            pass
    
    def create_widgets(self):
        title_label = tk.Label(self.root, text="银行交易明细分析工具 v3.0", 
                               font=("Microsoft YaHei", 16, "bold"), bg='#f0f0f0', fg="#333")
        title_label.pack(pady=10)
        
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(pady=10, padx=20, fill="both", expand=True)
        
        self.create_merge_tab()
        self.create_complete_tab()
        self.create_analysis_tab()
        self.create_behavior_tab()
        self.create_relation_tab()
        self.create_filter_tab()
        self.create_config_tab()
        
        self.status_label = tk.Label(self.root, text="✅ 就绪", 
                                     font=("Microsoft YaHei", 9), bg='#e0e0e0', fg="#666", 
                                     relief="sunken", anchor='w')
        self.status_label.pack(side="bottom", fill="x")
    # ================== 1. 数据合并模块 ==================
    def create_merge_tab(self):
        merge_frame = tk.Frame(self.notebook, bg='#f0f0f0')
        self.notebook.add(merge_frame, text="1. 📦 数据合并")
        
        left_frame = tk.Frame(merge_frame, bg='#f0f0f0')
        left_frame.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        
        right_frame = tk.Frame(merge_frame, bg='#f0f0f0')
        right_frame.pack(side="right", fill="both", expand=True, padx=10, pady=10)
        
        # ----- 操作说明 -----
        info_frame = tk.LabelFrame(left_frame, text="📖 操作说明", font=("Microsoft YaHei", 10, "bold"),
                                   bg='#f0f0f0', fg='#333')
        info_frame.pack(fill="x", pady=5)
        
        info_text = """【功能说明】
自动合并同一文件夹下所有 CSV 文件。

【识别规则】
- 文件名包含"交易明细" → 合并为交易明细文件
- 文件名包含"账户信息" → 合并为账户信息文件
- 文件名包含"合并结果" → 自动跳过

【操作步骤】
1. 点击"选择文件夹"选取 CSV 所在目录
2. 点击"开始合并"
3. 等待程序自动识别并合并

【输出结果】
- 银行交易明细合并结果_时间戳.xlsx
- 银行账户信息合并结果_时间戳.xlsx
- 保存在您选择的文件夹内"""
        
        info_label = tk.Label(info_frame, text=info_text, justify="left", bg='#f0f0f0', 
                             font=("Microsoft YaHei", 9), fg="#444")
        info_label.pack(padx=10, pady=10, anchor="w")
        
        # ----- 操作区域 -----
        path_frame = tk.LabelFrame(left_frame, text="📁 文件夹选择", font=("Microsoft YaHei", 11, "bold"),
                                   bg='#f0f0f0', fg='#333')
        path_frame.pack(fill="x", pady=5)
        
        tk.Label(path_frame, text="CSV文件夹:", bg='#f0f0f0', font=("Microsoft YaHei", 10)).pack(anchor="w", padx=10, pady=5)
        self.merge_folder_var = tk.StringVar()
        merge_folder_entry = tk.Entry(path_frame, textvariable=self.merge_folder_var, font=("Microsoft YaHei", 10))
        merge_folder_entry.pack(fill="x", padx=10, pady=5)
        
        btn_frame = tk.Frame(path_frame, bg='#f0f0f0')
        btn_frame.pack(pady=10)
        
        tk.Button(btn_frame, text="📁 选择文件夹", command=self.select_merge_folder,
                 font=("Microsoft YaHei", 10), bg="#4CAF50", fg="white", padx=15).pack(side="left", padx=5)
        
        self.merge_btn = tk.Button(left_frame, text="🚀 开始合并", command=self.start_merge,
                                  font=("Microsoft YaHei", 11, "bold"), bg="#2196F3", fg="white",
                                  padx=20, pady=5)
        self.merge_btn.pack(pady=10)
        
        self.merge_progress = ttk.Progressbar(left_frame, mode='indeterminate')
        self.merge_progress.pack(pady=10, fill="x")
        
        # ----- 日志区域 -----
        log_frame = tk.LabelFrame(right_frame, text="合并日志", font=("Microsoft YaHei", 10, "bold"),
                                  bg='#f0f0f0', fg='#333')
        log_frame.pack(fill="both", expand=True)
        
        self.merge_log_text = scrolledtext.ScrolledText(log_frame, height=25, 
                                                         font=("Consolas", 9),
                                                         bg='#2d2d2d', fg='#f8f8f2',
                                                         insertbackground='white')
        self.merge_log_text.pack(fill="both", expand=True, padx=5, pady=5)
    
    def select_merge_folder(self):
        folder = filedialog.askdirectory(title="选择包含CSV文件的文件夹")
        if folder:
            self.merge_folder_var.set(folder)
            self.log_merge(f"📂 已选择文件夹: {folder}")
    # ================== 2. 补全交易明细模块 ==================
    def create_complete_tab(self):
        complete_frame = tk.Frame(self.notebook, bg='#f0f0f0')
        self.notebook.add(complete_frame, text="2. 🔧 补全交易明细")
        
        left_frame = tk.Frame(complete_frame, bg='#f0f0f0')
        left_frame.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        
        right_frame = tk.Frame(complete_frame, bg='#f0f0f0')
        right_frame.pack(side="right", fill="both", expand=True, padx=10, pady=10)
        
        # ----- 操作说明 -----
        info_frame = tk.LabelFrame(left_frame, text="📖 操作说明", font=("Microsoft YaHei", 10, "bold"),
                                   bg='#f0f0f0', fg='#333')
        info_frame.pack(fill="x", pady=5)
        
        info_text = """【功能说明】
用账户信息补全交易明细中缺失的字段。

【补全内容】
- 交易卡号（通过交易账号匹配）
- 账户开户名称（通过账号/卡号匹配）
- 开户人证件号码（通过账号/卡号匹配）

【去重选项】
点击"开始补全数据"后会弹窗询问：
- 选择"是"：按交易流水号去重，保留第一条
- 选择"否"：保留所有记录（含重复）

【输入要求】
- 交易明细文件（Excel格式）
- 账户信息文件（Excel格式）

【输出结果】
- 交易明细_补全结果_时间戳.xlsx
- 保存在交易明细文件所在目录"""
        
        info_label = tk.Label(info_frame, text=info_text, justify="left", bg='#f0f0f0', 
                             font=("Microsoft YaHei", 9), fg="#444")
        info_label.pack(padx=10, pady=10, anchor="w")
        
        # ----- 文件选择 -----
        file_frame = tk.LabelFrame(left_frame, text="📁 文件选择", font=("Microsoft YaHei", 11, "bold"),
                                   bg='#f0f0f0', fg='#333')
        file_frame.pack(fill="x", pady=5)
        
        tk.Label(file_frame, text="交易明细文件:", bg='#f0f0f0', font=("Microsoft YaHei", 10)).pack(anchor="w", padx=10, pady=5)
        trans_file_frame = tk.Frame(file_frame, bg='#f0f0f0')
        trans_file_frame.pack(pady=5, padx=10, fill="x")
        
        self.trans_file_var = tk.StringVar(value=self.config["complete_config"].get("transaction_file", ""))
        trans_entry = tk.Entry(trans_file_frame, textvariable=self.trans_file_var, 
                               font=("Microsoft YaHei", 10), state="readonly",
                               bg='white', relief='solid', bd=1)
        trans_entry.pack(side="left", fill="x", expand=True, padx=(0, 10))
        tk.Button(trans_file_frame, text="选择", command=self.select_trans_file,
                 font=("Microsoft YaHei", 10), bg="#4CAF50", fg="white", width=6).pack(side="right")
        
        tk.Label(file_frame, text="账户信息文件:", bg='#f0f0f0', font=("Microsoft YaHei", 10)).pack(anchor="w", padx=10, pady=5)
        acc_file_frame = tk.Frame(file_frame, bg='#f0f0f0')
        acc_file_frame.pack(pady=5, padx=10, fill="x")
        
        self.acc_file_var = tk.StringVar(value=self.config["complete_config"].get("account_file", ""))
        acc_entry = tk.Entry(acc_file_frame, textvariable=self.acc_file_var, 
                             font=("Microsoft YaHei", 10), state="readonly",
                             bg='white', relief='solid', bd=1)
        acc_entry.pack(side="left", fill="x", expand=True, padx=(0, 10))
        tk.Button(acc_file_frame, text="选择", command=self.select_acc_file,
                 font=("Microsoft YaHei", 10), bg="#4CAF50", fg="white", width=6).pack(side="right")
        
        self.complete_btn = tk.Button(left_frame, text="🔧 开始补全数据", command=self.start_complete,
                                     font=("Microsoft YaHei", 11, "bold"), bg="#FF9800", fg="white",
                                     padx=20, pady=5)
        self.complete_btn.pack(pady=10)
        
        self.complete_progress = ttk.Progressbar(left_frame, mode='indeterminate')
        self.complete_progress.pack(pady=10, fill="x")
        
        # ----- 日志区域 -----
        log_frame = tk.LabelFrame(right_frame, text="补全日志", font=("Microsoft YaHei", 10, "bold"),
                                  bg='#f0f0f0', fg='#333')
        log_frame.pack(fill="both", expand=True)
        
        self.complete_log_text = scrolledtext.ScrolledText(log_frame, height=30, 
                                                            font=("Consolas", 9),
                                                            bg='#2d2d2d', fg='#f8f8f2',
                                                            insertbackground='white')
        self.complete_log_text.pack(fill="both", expand=True, padx=5, pady=5)
    # ================== 3. 基础分析模块 ==================
    def create_analysis_tab(self):
        analysis_frame = tk.Frame(self.notebook, bg='#f0f0f0')
        self.notebook.add(analysis_frame, text="3. 📊 基础分析")
        
        left_frame = tk.Frame(analysis_frame, bg='#f0f0f0')
        left_frame.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        
        right_frame = tk.Frame(analysis_frame, bg='#f0f0f0')
        right_frame.pack(side="right", fill="both", expand=True, padx=10, pady=10)
        
        # ----- 操作说明 -----
        info_frame = tk.LabelFrame(left_frame, text="📖 操作说明", font=("Microsoft YaHei", 10, "bold"),
                                   bg='#f0f0f0', fg='#333')
        info_frame.pack(fill="x", pady=5)
        
        info_text = """【功能说明】
对交易明细进行基础统计分析。

【分析内容】
1. 夜间交易标记（可自定义时间段）
2. 整数倍金额交易标记（可自定义最小金额）

【参数说明】
- 夜间交易时间段：默认 21:00-05:00（可调）
- 整数倍金额：默认 500 元以上（可调）

【输出结果】
- 1_夜间转账统计_时间戳.xlsx
- 2_统计分析报告_时间戳.xlsx
- 保存在分析结果_时间戳/文件夹中"""
        
        info_label = tk.Label(info_frame, text=info_text, justify="left", bg='#f0f0f0', 
                             font=("Microsoft YaHei", 9), fg="#444")
        info_label.pack(padx=10, pady=10, anchor="w")
        
        # ----- 文件选择 -----
        file_frame = tk.LabelFrame(left_frame, text="📁 文件选择", font=("Microsoft YaHei", 11, "bold"),
                                   bg='#f0f0f0', fg='#333')
        file_frame.pack(fill="x", pady=5)
        
        file_select_frame = tk.Frame(file_frame, bg='#f0f0f0')
        file_select_frame.pack(pady=10, padx=10, fill="x")
        
        self.file_path_var = tk.StringVar()
        self.file_entry = tk.Entry(file_select_frame, textvariable=self.file_path_var, 
                                   font=("Microsoft YaHei", 10), state="readonly",
                                   bg='white', relief='solid', bd=1)
        self.file_entry.pack(side="left", fill="x", expand=True, padx=(0, 10))
        self.select_btn = tk.Button(file_select_frame, text="📁 选择文件", command=self.select_file,
                                    font=("Microsoft YaHei", 10, "bold"), bg="#4CAF50", fg="white",
                                    padx=20, relief='flat', cursor='hand2')
        self.select_btn.pack(side="right")
        
        # ----- 参数设置 -----
        param_frame = tk.LabelFrame(left_frame, text="⚙️ 参数设置", font=("Microsoft YaHei", 11, "bold"),
                                    bg='#f0f0f0', fg='#333')
        param_frame.pack(fill="x", pady=5)
        
        time_frame = tk.Frame(param_frame, bg='#f0f0f0')
        time_frame.pack(pady=10, padx=10, fill="x")
        
        tk.Label(time_frame, text="夜间交易时间段:", font=("Microsoft YaHei", 10), 
                bg='#f0f0f0').pack(side="left")
        
        self.night_start = tk.StringVar(value=str(self.config["analysis_params"]["night_start"]))
        self.night_end = tk.StringVar(value=str(self.config["analysis_params"]["night_end"]))
        
        tk.Spinbox(time_frame, from_=0, to=23, width=5, 
                  textvariable=self.night_start, font=("Microsoft YaHei", 10)).pack(side="left", padx=5)
        tk.Label(time_frame, text=":00 至", font=("Microsoft YaHei", 10), 
                bg='#f0f0f0').pack(side="left")
        tk.Spinbox(time_frame, from_=0, to=23, width=5, 
                  textvariable=self.night_end, font=("Microsoft YaHei", 10)).pack(side="left", padx=5)
        tk.Label(time_frame, text=":00", font=("Microsoft YaHei", 10), 
                bg='#f0f0f0').pack(side="left")
        
        amount_frame = tk.Frame(param_frame, bg='#f0f0f0')
        amount_frame.pack(pady=10, padx=10, fill="x")
        
        tk.Label(amount_frame, text="整数倍交易最小金额:", font=("Microsoft YaHei", 10), 
                bg='#f0f0f0').pack(side="left")
        
        self.min_amount = tk.StringVar(value=str(self.config["analysis_params"]["min_amount"]))
        tk.Spinbox(amount_frame, from_=100, to=10000, width=8, 
                  textvariable=self.min_amount, font=("Microsoft YaHei", 10)).pack(side="left", padx=5)
        tk.Label(amount_frame, text="元 (100的整数倍)", font=("Microsoft YaHei", 9), 
                bg='#f0f0f0', fg='#999').pack(side="left")
        
        self.analyze_btn = tk.Button(left_frame, text="🚀 开始分析", command=self.start_analysis,
                                    font=("Microsoft YaHei", 12, "bold"), bg="#2196F3", fg="white",
                                    height=2, width=20, relief='flat', cursor='hand2')
        self.analyze_btn.pack(pady=15)
        
        self.progress = ttk.Progressbar(left_frame, mode='indeterminate')
        self.progress.pack(pady=10, fill="x")
        
        # ----- 日志区域 -----
        log_frame = tk.LabelFrame(right_frame, text="分析日志", font=("Microsoft YaHei", 10, "bold"),
                                  bg='#f0f0f0', fg='#333')
        log_frame.pack(fill="both", expand=True)
        
        self.log_text = scrolledtext.ScrolledText(log_frame, height=20, 
                                                   font=("Consolas", 9),
                                                   bg='#2d2d2d', fg='#f8f8f2',
                                                   insertbackground='white')
        self.log_text.pack(fill="both", expand=True, padx=5, pady=5)
    # ================== 4. 账户行为分析 ==================
    def create_behavior_tab(self):
        behavior_frame = tk.Frame(self.notebook, bg='#f0f0f0')
        self.notebook.add(behavior_frame, text="4. 📈 账户行为分析")
        
        left_frame = tk.Frame(behavior_frame, bg='#f0f0f0')
        left_frame.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        
        right_frame = tk.Frame(behavior_frame, bg='#f0f0f0')
        right_frame.pack(side="right", fill="both", expand=True, padx=10, pady=10)
        
        # ----- 操作说明 -----
        info_frame = tk.LabelFrame(left_frame, text="📖 操作说明", font=("Microsoft YaHei", 10, "bold"),
                                   bg='#f0f0f0', fg='#333')
        info_frame.pack(fill="x", pady=5)
        
        info_text = """【功能说明】
识别高频交易账户和大额交易账户。

【分析规则】
- 高频账户：交易次数 ≥ 设定阈值
- 大额账户：交易总额 ≥ 设定阈值
- 时间窗口：分析最近N天的活跃情况

【参数说明】
- 高频交易阈值：默认 10 次
- 大额交易阈值：默认 50,000 元
- 时间窗口：默认 30 天

【输出结果】
- 账户行为分析_时间戳.xlsx
- 包含：高频账户、大额账户、账户汇总三个Sheet"""
        
        info_label = tk.Label(info_frame, text=info_text, justify="left", bg='#f0f0f0', 
                             font=("Microsoft YaHei", 9), fg="#444")
        info_label.pack(padx=10, pady=10, anchor="w")
        
        # ----- 文件选择 -----
        file_frame = tk.LabelFrame(left_frame, text="📁 文件选择", font=("Microsoft YaHei", 11, "bold"),
                                   bg='#f0f0f0', fg='#333')
        file_frame.pack(fill="x", pady=5)
        
        file_select_frame = tk.Frame(file_frame, bg='#f0f0f0')
        file_select_frame.pack(pady=10, padx=10, fill="x")
        
        self.behavior_file_var = tk.StringVar()
        behavior_file_entry = tk.Entry(file_select_frame, textvariable=self.behavior_file_var, 
                                       font=("Microsoft YaHei", 10), state="readonly",
                                       bg='white', relief='solid', bd=1)
        behavior_file_entry.pack(side="left", fill="x", expand=True, padx=(0, 10))
        tk.Button(file_select_frame, text="选择文件", command=self.select_behavior_file,
                 font=("Microsoft YaHei", 10), bg="#4CAF50", fg="white").pack(side="right")
        
        # ----- 参数设置 -----
        param_frame = tk.LabelFrame(left_frame, text="⚙️ 参数设置", font=("Microsoft YaHei", 11, "bold"),
                                    bg='#f0f0f0', fg='#333')
        param_frame.pack(fill="x", pady=5)
        
        tk.Label(param_frame, text="高频交易阈值(次数):", bg='#f0f0f0', font=("Microsoft YaHei", 10)).pack(anchor="w", padx=10, pady=5)
        self.freq_threshold = tk.StringVar(value=str(self.config["behavior_analysis"]["frequency_threshold"]))
        tk.Entry(param_frame, textvariable=self.freq_threshold, width=10, font=("Microsoft YaHei", 10)).pack(anchor="w", padx=10)
        
        tk.Label(param_frame, text="大额交易阈值(元):", bg='#f0f0f0', font=("Microsoft YaHei", 10)).pack(anchor="w", padx=10, pady=5)
        self.amount_threshold = tk.StringVar(value=str(self.config["behavior_analysis"]["amount_threshold"]))
        tk.Entry(param_frame, textvariable=self.amount_threshold, width=15, font=("Microsoft YaHei", 10)).pack(anchor="w", padx=10)
        
        tk.Label(param_frame, text="时间窗口(天):", bg='#f0f0f0', font=("Microsoft YaHei", 10)).pack(anchor="w", padx=10, pady=5)
        self.time_window = tk.StringVar(value=str(self.config["behavior_analysis"]["time_window_days"]))
        tk.Entry(param_frame, textvariable=self.time_window, width=10, font=("Microsoft YaHei", 10)).pack(anchor="w", padx=10)
        
        self.behavior_btn = tk.Button(left_frame, text="🔍 开始行为分析", command=self.start_behavior_analysis,
                                     font=("Microsoft YaHei", 11, "bold"), bg="#9C27B0", fg="white",
                                     padx=20, pady=5)
        self.behavior_btn.pack(pady=10)
        
        self.behavior_progress = ttk.Progressbar(left_frame, mode='indeterminate')
        self.behavior_progress.pack(pady=10, fill="x")
        
        # ----- 结果区域 -----
        result_frame = tk.LabelFrame(right_frame, text="行为分析结果", font=("Microsoft YaHei", 10, "bold"),
                                     bg='#f0f0f0', fg='#333')
        result_frame.pack(fill="both", expand=True)
        
        self.behavior_result_text = scrolledtext.ScrolledText(result_frame, height=25, 
                                                               font=("Consolas", 9),
                                                               bg='#2d2d2d', fg='#f8f8f2',
                                                               insertbackground='white')
        self.behavior_result_text.pack(fill="both", expand=True, padx=5, pady=5)
    # ================== 5. 关联分析 ==================
    def create_relation_tab(self):
        relation_frame = tk.Frame(self.notebook, bg='#f0f0f0')
        self.notebook.add(relation_frame, text="5. 🔗 关联分析")
        
        left_frame = tk.Frame(relation_frame, bg='#f0f0f0')
        left_frame.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        
        right_frame = tk.Frame(relation_frame, bg='#f0f0f0')
        right_frame.pack(side="right", fill="both", expand=True, padx=10, pady=10)
        
        # ----- 操作说明 -----
        info_frame = tk.LabelFrame(left_frame, text="📖 操作说明", font=("Microsoft YaHei", 10, "bold"),
                                   bg='#f0f0f0', fg='#333')
        info_frame.pack(fill="x", pady=5)
        
        info_text = """【功能说明】
发现账户之间的交易关联关系。

【分析规则】
- 基于交易账号和交易方账号构建关联网络
- 统计每对账户间的交易次数和总金额
- 筛选符合条件的显著关联关系

【参数说明】
- 最小交易次数：默认 3 次
- 最小总金额：默认 10,000 元
- 分析类型：收入关系/支出关系/双向关系

【输出结果】
- 关联分析结果_时间戳.xlsx
- 包含：来源账户、目标账户、交易次数、总金额"""
        
        info_label = tk.Label(info_frame, text=info_text, justify="left", bg='#f0f0f0', 
                             font=("Microsoft YaHei", 9), fg="#444")
        info_label.pack(padx=10, pady=10, anchor="w")
        
        # ----- 文件选择 -----
        file_frame = tk.LabelFrame(left_frame, text="📁 文件选择", font=("Microsoft YaHei", 11, "bold"),
                                   bg='#f0f0f0', fg='#333')
        file_frame.pack(fill="x", pady=5)
        
        file_select_frame = tk.Frame(file_frame, bg='#f0f0f0')
        file_select_frame.pack(pady=10, padx=10, fill="x")
        
        self.relation_file_var = tk.StringVar()
        relation_file_entry = tk.Entry(file_select_frame, textvariable=self.relation_file_var, 
                                       font=("Microsoft YaHei", 10), state="readonly",
                                       bg='white', relief='solid', bd=1)
        relation_file_entry.pack(side="left", fill="x", expand=True, padx=(0, 10))
        tk.Button(file_select_frame, text="选择文件", command=self.select_relation_file,
                 font=("Microsoft YaHei", 10), bg="#4CAF50", fg="white").pack(side="right")
        
        # ----- 参数设置 -----
        param_frame = tk.LabelFrame(left_frame, text="⚙️ 参数设置", font=("Microsoft YaHei", 11, "bold"),
                                    bg='#f0f0f0', fg='#333')
        param_frame.pack(fill="x", pady=5)
        
        tk.Label(param_frame, text="最小交易次数:", bg='#f0f0f0', font=("Microsoft YaHei", 10)).pack(anchor="w", padx=10, pady=5)
        self.min_tx_count = tk.StringVar(value="3")
        tk.Entry(param_frame, textvariable=self.min_tx_count, width=10, font=("Microsoft YaHei", 10)).pack(anchor="w", padx=10)
        
        tk.Label(param_frame, text="最小总金额(元):", bg='#f0f0f0', font=("Microsoft YaHei", 10)).pack(anchor="w", padx=10, pady=5)
        self.min_total_amount = tk.StringVar(value="10000")
        tk.Entry(param_frame, textvariable=self.min_total_amount, width=15, font=("Microsoft YaHei", 10)).pack(anchor="w", padx=10)
        
        tk.Label(param_frame, text="分析类型:", bg='#f0f0f0', font=("Microsoft YaHei", 10)).pack(anchor="w", padx=10, pady=5)
        self.relation_type = tk.StringVar(value="both")
        tk.Radiobutton(param_frame, text="收入关系", variable=self.relation_type, 
                      value="income", bg='#f0f0f0', font=("Microsoft YaHei", 10)).pack(anchor="w", padx=20)
        tk.Radiobutton(param_frame, text="支出关系", variable=self.relation_type, 
                      value="outcome", bg='#f0f0f0', font=("Microsoft YaHei", 10)).pack(anchor="w", padx=20)
        tk.Radiobutton(param_frame, text="双向关系", variable=self.relation_type, 
                      value="both", bg='#f0f0f0', font=("Microsoft YaHei", 10)).pack(anchor="w", padx=20)
        
        self.relation_btn = tk.Button(left_frame, text="🔍 开始关联分析", command=self.start_relation_analysis,
                                     font=("Microsoft YaHei", 11, "bold"), bg="#E91E63", fg="white",
                                     padx=20, pady=5)
        self.relation_btn.pack(pady=10)
        
        self.relation_progress = ttk.Progressbar(left_frame, mode='indeterminate')
        self.relation_progress.pack(pady=10, fill="x")
        
        # ----- 结果区域 -----
        result_frame = tk.LabelFrame(right_frame, text="关联分析结果", font=("Microsoft YaHei", 10, "bold"),
                                     bg='#f0f0f0', fg='#333')
        result_frame.pack(fill="both", expand=True)
        
        self.relation_result_text = scrolledtext.ScrolledText(result_frame, height=25, 
                                                               font=("Consolas", 9),
                                                               bg='#2d2d2d', fg='#f8f8f2',
                                                               insertbackground='white')
        self.relation_result_text.pack(fill="both", expand=True, padx=5, pady=5)
    # ================== 6. 高级筛选 ==================
    def create_filter_tab(self):
        filter_frame = tk.Frame(self.notebook, bg='#f0f0f0')
        self.notebook.add(filter_frame, text="6. 🔍 高级筛选")
        
        left_frame = tk.Frame(filter_frame, bg='#f0f0f0')
        left_frame.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        
        right_frame = tk.Frame(filter_frame, bg='#f0f0f0')
        right_frame.pack(side="right", fill="both", expand=True, padx=10, pady=10)
        
        # ----- 操作说明 -----
        info_frame = tk.LabelFrame(left_frame, text="📖 操作说明", font=("Microsoft YaHei", 10, "bold"),
                                   bg='#f0f0f0', fg='#333')
        info_frame.pack(fill="x", pady=5)
        
        info_text = """【功能说明】
多条件组合筛选交易明细数据。

【筛选条件】
- 金额范围：最小金额 ~ 最大金额
- 时间范围：开始日期 ~ 结束日期
- 交易方向：全部/收入/支出
- 关键词：按户名或摘要搜索

【操作说明】
- 所有条件可单独使用或组合使用
- 条件留空表示不限制该条件
- 筛选结果可直接导出为Excel文件

【输出结果】
- 筛选结果_时间戳.xlsx"""
        
        info_label = tk.Label(info_frame, text=info_text, justify="left", bg='#f0f0f0', 
                             font=("Microsoft YaHei", 9), fg="#444")
        info_label.pack(padx=10, pady=10, anchor="w")
        
        # ----- 文件选择 -----
        file_frame = tk.LabelFrame(left_frame, text="📁 文件选择", font=("Microsoft YaHei", 11, "bold"),
                                   bg='#f0f0f0', fg='#333')
        file_frame.pack(fill="x", pady=5)
        
        file_select_frame = tk.Frame(file_frame, bg='#f0f0f0')
        file_select_frame.pack(pady=10, padx=10, fill="x")
        
        self.filter_file_var = tk.StringVar()
        filter_file_entry = tk.Entry(file_select_frame, textvariable=self.filter_file_var, 
                                     font=("Microsoft YaHei", 10), state="readonly",
                                     bg='white', relief='solid', bd=1)
        filter_file_entry.pack(side="left", fill="x", expand=True, padx=(0, 10))
        tk.Button(file_select_frame, text="选择文件", command=self.select_filter_file,
                 font=("Microsoft YaHei", 10), bg="#4CAF50", fg="white").pack(side="right")
        
        # ----- 筛选条件 -----
        condition_frame = tk.LabelFrame(left_frame, text="🔍 筛选条件", font=("Microsoft YaHei", 11, "bold"),
                                        bg='#f0f0f0', fg='#333')
        condition_frame.pack(fill="x", pady=5)
        
        tk.Label(condition_frame, text="金额范围(元):", bg='#f0f0f0', font=("Microsoft YaHei", 10)).pack(anchor="w", padx=10, pady=5)
        amount_range_frame = tk.Frame(condition_frame, bg='#f0f0f0')
        amount_range_frame.pack(anchor="w", padx=10)
        
        self.amount_min = tk.StringVar()
        self.amount_max = tk.StringVar()
        tk.Entry(amount_range_frame, textvariable=self.amount_min, width=15, font=("Microsoft YaHei", 10)).pack(side="left", padx=5)
        tk.Label(amount_range_frame, text="至", bg='#f0f0f0', font=("Microsoft YaHei", 10)).pack(side="left")
        tk.Entry(amount_range_frame, textvariable=self.amount_max, width=15, font=("Microsoft YaHei", 10)).pack(side="left", padx=5)
        
        tk.Label(condition_frame, text="时间范围:", bg='#f0f0f0', font=("Microsoft YaHei", 10)).pack(anchor="w", padx=10, pady=5)
        time_range_frame = tk.Frame(condition_frame, bg='#f0f0f0')
        time_range_frame.pack(anchor="w", padx=10)
        
        self.date_start = tk.StringVar()
        self.date_end = tk.StringVar()
        tk.Entry(time_range_frame, textvariable=self.date_start, width=20, font=("Microsoft YaHei", 10)).pack(side="left", padx=5)
        tk.Label(time_range_frame, text="至", bg='#f0f0f0', font=("Microsoft YaHei", 10)).pack(side="left")
        tk.Entry(time_range_frame, textvariable=self.date_end, width=20, font=("Microsoft YaHei", 10)).pack(side="left", padx=5)
        
        tk.Label(condition_frame, text="交易方向:", bg='#f0f0f0', font=("Microsoft YaHei", 10)).pack(anchor="w", padx=10, pady=5)
        direction_frame = tk.Frame(condition_frame, bg='#f0f0f0')
        direction_frame.pack(anchor="w", padx=10)
        self.tx_direction = tk.StringVar(value="all")
        tk.Radiobutton(direction_frame, text="全部", variable=self.tx_direction, 
                      value="all", bg='#f0f0f0', font=("Microsoft YaHei", 10)).pack(side="left", padx=5)
        tk.Radiobutton(direction_frame, text="收入", variable=self.tx_direction, 
                      value="income", bg='#f0f0f0', font=("Microsoft YaHei", 10)).pack(side="left", padx=5)
        tk.Radiobutton(direction_frame, text="支出", variable=self.tx_direction, 
                      value="outcome", bg='#f0f0f0', font=("Microsoft YaHei", 10)).pack(side="left", padx=5)
        
        tk.Label(condition_frame, text="关键词搜索(户名/摘要):", bg='#f0f0f0', font=("Microsoft YaHei", 10)).pack(anchor="w", padx=10, pady=5)
        self.keyword = tk.StringVar()
        tk.Entry(condition_frame, textvariable=self.keyword, width=40, font=("Microsoft YaHei", 10)).pack(anchor="w", padx=10)
        
        btn_frame = tk.Frame(left_frame, bg='#f0f0f0')
        btn_frame.pack(pady=10)
        
        self.filter_btn = tk.Button(btn_frame, text="🔍 执行筛选", command=self.start_filter,
                                   font=("Microsoft YaHei", 11, "bold"), bg="#FF9800", fg="white",
                                   padx=20, pady=5)
        self.filter_btn.pack(side="left", padx=5)
        
        self.clear_btn = tk.Button(btn_frame, text="🗑️ 清空条件", command=self.clear_filter_conditions,
                                  font=("Microsoft YaHei", 11, "bold"), bg="#9E9E9E", fg="white",
                                  padx=20, pady=5)
        self.clear_btn.pack(side="left", padx=5)
        
        self.filter_progress = ttk.Progressbar(left_frame, mode='indeterminate')
        self.filter_progress.pack(pady=10, fill="x")
        
        # ----- 结果区域 -----
        result_frame = tk.LabelFrame(right_frame, text="筛选结果", font=("Microsoft YaHei", 10, "bold"),
                                     bg='#f0f0f0', fg='#333')
        result_frame.pack(fill="both", expand=True)
        
        self.filter_result_text = scrolledtext.ScrolledText(result_frame, height=25, 
                                                             font=("Consolas", 9),
                                                             bg='#2d2d2d', fg='#f8f8f2',
                                                             insertbackground='white')
        self.filter_result_text.pack(fill="both", expand=True, padx=5, pady=5)
        
        self.filter_stats_label = tk.Label(right_frame, text="", font=("Microsoft YaHei", 9),
                                           bg='#f0f0f0', fg='#666')
        self.filter_stats_label.pack(pady=2)
    # ================== 7. 配置管理 ==================
    def create_config_tab(self):
        config_frame = tk.Frame(self.notebook, bg='#f0f0f0')
        self.notebook.add(config_frame, text="7. ⚙️ 配置管理")
        
        left_frame = tk.Frame(config_frame, bg='#f0f0f0')
        left_frame.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        
        right_frame = tk.Frame(config_frame, bg='#f0f0f0')
        right_frame.pack(side="right", fill="both", expand=True, padx=10, pady=10)
        
        # ----- 操作说明 -----
        info_frame = tk.LabelFrame(left_frame, text="📖 操作说明", font=("Microsoft YaHei", 10, "bold"),
                                   bg='#f0f0f0', fg='#333')
        info_frame.pack(fill="x", pady=5)
        
        info_text = """【功能说明】
自定义列名映射和输出列顺序。

【列名映射】
将Excel中的实际列名映射到程序识别的标准字段。
- 左侧为程序标准字段名
- 右侧输入框中填入Excel中的实际列名

【输出顺序】
设置导出Excel时各列的排列顺序。
- 每行写一个列名
- 按顺序从上到下排列

【操作步骤】
1. 修改映射或顺序后点击"保存"
2. 配置自动保存，下次启动自动加载"""
        
        info_label = tk.Label(info_frame, text=info_text, justify="left", bg='#f0f0f0', 
                             font=("Microsoft YaHei", 9), fg="#444")
        info_label.pack(padx=10, pady=10, anchor="w")
        
        # ----- 列名映射 -----
        mapping_frame = tk.LabelFrame(left_frame, text="列名映射配置", font=("Microsoft YaHei", 11, "bold"),
                                      bg='#f0f0f0', fg='#333')
        mapping_frame.pack(fill="both", expand=True, pady=5)
        
        self.mapping_entries = {}
        standard_fields = ["交易时间", "交易金额", "交易卡号", "交易账号", "账户开户名称", 
                          "开户人证件号码", "交易方户名", "交易方账号", "交易方证件号码", 
                          "收付标志", "交易摘要", "查询反馈结果原因", "交易流水号"]
        
        canvas = tk.Canvas(mapping_frame, bg='#f0f0f0', height=200)
        scrollbar = ttk.Scrollbar(mapping_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='#f0f0f0')
        
        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        for i, field in enumerate(standard_fields):
            field_frame = tk.Frame(scrollable_frame, bg='#f0f0f0', pady=3)
            field_frame.pack(fill="x", padx=5)
            
            tk.Label(field_frame, text=f"{field}:", width=18, anchor='w',
                    font=("Microsoft YaHei", 10), bg='#f0f0f0').pack(side="left")
            
            entry_var = tk.StringVar(value=self.config["column_mapping"].get(field, field))
            entry = tk.Entry(field_frame, textvariable=entry_var, 
                           font=("Microsoft YaHei", 10), width=25)
            entry.pack(side="left", padx=5)
            
            self.mapping_entries[field] = entry_var
        
        tk.Button(mapping_frame, text="💾 保存映射配置", command=self.save_mapping,
                 font=("Microsoft YaHei", 10, "bold"), bg="#FF9800", fg="white",
                 pady=5, padx=20).pack(pady=10)
        
        # ----- 输出顺序 -----
        order_frame = tk.LabelFrame(right_frame, text="输出列顺序", font=("Microsoft YaHei", 11, "bold"),
                                    bg='#f0f0f0', fg='#333')
        order_frame.pack(fill="both", expand=True, pady=5)
        
        self.order_text = tk.Text(order_frame, height=15, font=("Consolas", 10))
        self.order_text.pack(padx=10, pady=10, fill="both", expand=True)
        self.order_text.insert("1.0", "\n".join(self.config["output_order"]))
        
        tk.Button(order_frame, text="💾 保存列顺序", command=self.save_order,
                 font=("Microsoft YaHei", 10, "bold"), bg="#FF9800", fg="white",
                 pady=5, padx=20).pack(pady=10)
        
        # ----- 重置 -----
        tk.Button(right_frame, text="🔄 重置所有配置", command=self.reset_config,
                 font=("Microsoft YaHei", 10, "bold"), bg="#f44336", fg="white",
                 pady=5, padx=20).pack(pady=10)
    # ================== 辅助功能 ==================
    def select_trans_file(self):
        file_path = filedialog.askopenfilename(
            title="选择交易明细Excel文件",
            filetypes=[("Excel文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
        )
        if file_path:
            self.trans_file_var.set(file_path)
            self.config["complete_config"]["transaction_file"] = file_path
            self.save_config()
            self.log_complete(f"📂 已选择交易明细: {os.path.basename(file_path)}")
    
    def select_acc_file(self):
        file_path = filedialog.askopenfilename(
            title="选择账户信息Excel文件",
            filetypes=[("Excel文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
        )
        if file_path:
            self.acc_file_var.set(file_path)
            self.config["complete_config"]["account_file"] = file_path
            self.save_config()
            self.log_complete(f"📂 已选择账户信息: {os.path.basename(file_path)}")
    
    def select_file(self):
        file_path = filedialog.askopenfilename(
            title="选择交易明细Excel文件",
            filetypes=[("Excel文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
        )
        if file_path:
            self.file_path_var.set(file_path)
            self.current_file = file_path
            self.log_analysis(f"📂 已选择文件: {os.path.basename(file_path)}")
    
    def select_behavior_file(self):
        file_path = filedialog.askopenfilename(
            title="选择交易明细Excel文件",
            filetypes=[("Excel文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
        )
        if file_path:
            self.behavior_file_var.set(file_path)
            self.log_behavior(f"📂 已选择文件: {os.path.basename(file_path)}")
    
    def select_relation_file(self):
        file_path = filedialog.askopenfilename(
            title="选择交易明细Excel文件",
            filetypes=[("Excel文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
        )
        if file_path:
            self.relation_file_var.set(file_path)
            self.log_relation(f"📂 已选择文件: {os.path.basename(file_path)}")
    
    def select_filter_file(self):
        file_path = filedialog.askopenfilename(
            title="选择交易明细Excel文件",
            filetypes=[("Excel文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
        )
        if file_path:
            self.filter_file_var.set(file_path)
            self.log_filter(f"📂 已选择文件: {os.path.basename(file_path)}")
    
    def save_mapping(self):
        for field, var in self.mapping_entries.items():
            self.config["column_mapping"][field] = var.get()
        self.save_config()
        messagebox.showinfo("成功", "列映射配置已保存！")
    
    def save_order(self):
        order_text = self.order_text.get("1.0", tk.END).strip()
        order_list = [line.strip() for line in order_text.split('\n') if line.strip()]
        if order_list:
            self.config["output_order"] = order_list
            self.save_config()
            messagebox.showinfo("成功", "输出列顺序已保存！")
    
    def reset_config(self):
        if messagebox.askyesno("确认", "这将重置所有配置为默认值，确定吗？"):
            if os.path.exists(self.config_file):
                os.remove(self.config_file)
            self.load_config()
            messagebox.showinfo("成功", "配置已重置为默认值！")
    
    def clear_filter_conditions(self):
        """清空所有筛选条件"""
        self.amount_min.set("")
        self.amount_max.set("")
        self.date_start.set("")
        self.date_end.set("")
        self.tx_direction.set("all")
        self.keyword.set("")
        self.log_filter("✅ 已清空所有筛选条件")
        self.filter_stats_label.config(text="")
    # ================== 日志功能 ==================
    def log_merge(self, message):
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.merge_log_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.merge_log_text.see(tk.END)
        self.root.update()
    
    def log_complete(self, message):
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.complete_log_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.complete_log_text.see(tk.END)
        self.root.update()
    
    def log_analysis(self, message):
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.log_text.see(tk.END)
        self.root.update()
    
    def log_behavior(self, message):
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.behavior_result_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.behavior_result_text.see(tk.END)
        self.root.update()
    
    def log_relation(self, message):
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.relation_result_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.relation_result_text.see(tk.END)
        self.root.update()
    
    def log_filter(self, message):
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.filter_result_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.filter_result_text.see(tk.END)
        self.root.update()
    # ================== 合并功能 ==================
    def clean_text(self, text):
        if pd.isna(text):
            return ""
        text = str(text)
        for char in ['\t', '\n', '\r', '\u3000']:
            text = text.replace(char, '')
        return text.strip()
    
    def remove_scientific(self, value):
        if not value or value == "":
            return value
        try:
            if 'e+' in str(value).lower() or 'e-' in str(value).lower():
                return format(Decimal(str(value)), 'f')
            return str(value)
        except:
            return str(value)
    
    def read_csv_with_encoding(self, file_path, fallback_encoding):
        encodings = ['utf-8', fallback_encoding, 'gb2312', 'gb18030', 'latin1']
        for encoding in encodings:
            try:
                df = pd.read_csv(file_path, dtype=str, encoding=encoding, on_bad_lines='skip')
                return df
            except:
                continue
        raise Exception(f"无法读取文件: {file_path}")
    
    def start_merge(self):
        if not self.merge_folder_var.get():
            messagebox.showwarning("警告", "请先选择CSV文件夹！")
            return
        if self.merge_in_progress:
            messagebox.showwarning("警告", "合并任务正在进行中！")
            return
        thread = threading.Thread(target=self.run_merge)
        thread.daemon = True
        thread.start()
    
    def run_merge(self):
        try:
            self.merge_in_progress = True
            self.merge_btn.config(state="disabled", text="⏳ 合并中...")
            self.merge_progress.start()
            
            self.log_merge("="*60)
            self.log_merge("开始合并CSV文件")
            self.log_merge("="*60)
            
            folder_path = self.merge_folder_var.get()
            self.log_merge(f"📁 扫描文件夹: {folder_path}")
            
            all_csv_files = glob.glob(os.path.join(folder_path, "*.csv"))
            self.log_merge(f"📄 发现 {len(all_csv_files)} 个CSV文件")
            
            # 分类文件
            trans_files = []
            acc_files = []
            skipped_files = []
            
            for file in all_csv_files:
                filename = os.path.basename(file)
                if "合并结果" in filename:
                    skipped_files.append(filename)
                    continue
                elif "交易明细" in filename:
                    trans_files.append(file)
                elif "账户信息" in filename:
                    acc_files.append(file)
                else:
                    skipped_files.append(filename)
            
            self.log_merge(f"\n📊 分类结果:")
            self.log_merge(f"  交易明细文件: {len(trans_files)} 个")
            self.log_merge(f"  账户信息文件: {len(acc_files)} 个")
            self.log_merge(f"  跳过文件: {len(skipped_files)} 个")
            
            if skipped_files:
                self.log_merge(f"  (已跳过: {', '.join(skipped_files[:5])}{'...' if len(skipped_files)>5 else ''})")
            
            results = {}
            
            # 合并交易明细
            if trans_files:
                self.log_merge(f"\n📥 合并交易明细...")
                all_data = []
                for file in trans_files:
                    try:
                        df = self.read_csv_with_encoding(file, 'gb18030')
                        df.columns = [self.clean_text(c) for c in df.columns]
                        for col in df.columns:
                            df[col] = df[col].apply(self.clean_text).apply(self.remove_scientific)
                        df['来源文件'] = os.path.basename(file)
                        all_data.append(df)
                        self.log_merge(f"  ✓ {os.path.basename(file)}: {len(df)}行")
                    except Exception as e:
                        self.log_merge(f"  ✗ {os.path.basename(file)}: {e}")
                
                if all_data:
                    merged_df = pd.concat(all_data, ignore_index=True)
                    for col in merged_df.columns:
                        merged_df[col] = merged_df[col].apply(self.clean_text)
                    
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    output_filename = f"银行交易明细合并结果_{timestamp}.xlsx"
                    output_file = os.path.join(folder_path, output_filename)
                    
                    merged_df.to_excel(output_file, index=False)
                    wb = load_workbook(output_file)
                    for row in wb.active.iter_rows():
                        for cell in row:
                            cell.number_format = "@"
                    wb.save(output_file)
                    
                    self.log_merge(f"  ✅ 输出: {output_filename} ({len(merged_df)}行)")
                    results['交易明细'] = output_file
            
            # 合并账户信息
            if acc_files:
                self.log_merge(f"\n📥 合并账户信息...")
                all_data = []
                for file in acc_files:
                    try:
                        df = self.read_csv_with_encoding(file, 'gbk')
                        df.columns = [self.clean_text(c) for c in df.columns]
                        for col in df.columns:
                            df[col] = df[col].apply(self.clean_text).apply(self.remove_scientific)
                        df['来源文件'] = os.path.basename(file)
                        all_data.append(df)
                        self.log_merge(f"  ✓ {os.path.basename(file)}: {len(df)}行")
                    except Exception as e:
                        self.log_merge(f"  ✗ {os.path.basename(file)}: {e}")
                
                if all_data:
                    merged_df = pd.concat(all_data, ignore_index=True)
                    for col in merged_df.columns:
                        merged_df[col] = merged_df[col].apply(self.clean_text)
                    
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    output_filename = f"银行账户信息合并结果_{timestamp}.xlsx"
                    output_file = os.path.join(folder_path, output_filename)
                    
                    merged_df.to_excel(output_file, index=False)
                    wb = load_workbook(output_file)
                    for row in wb.active.iter_rows():
                        for cell in row:
                            cell.number_format = "@"
                    wb.save(output_file)
                    
                    self.log_merge(f"  ✅ 输出: {output_filename} ({len(merged_df)}行)")
                    results['账户信息'] = output_file
            
            self.log_merge("\n" + "="*60)
            self.log_merge(f"合并完成！成功处理 {len(results)} 个类别")
            for key, val in results.items():
                self.log_merge(f"  ✅ {key}: {os.path.basename(val)}")
            messagebox.showinfo("完成", f"合并完成！\n交易明细: {len(trans_files)}个文件\n账户信息: {len(acc_files)}个文件\n\n结果保存在:\n{folder_path}")
            
        except Exception as e:
            self.log_merge(f"❌ 合并失败: {e}")
            import traceback
            self.log_merge(traceback.format_exc())
            messagebox.showerror("错误", f"合并失败:\n{e}")
        finally:
            self.merge_progress.stop()
            self.merge_btn.config(state="normal", text="🚀 开始合并")
            self.merge_in_progress = False
    # ================== 补全功能 ==================
    def clean_dataframe(self, df):
        df = df.copy()
        for col in df.columns:
            if df[col].dtype == 'object':
                df[col] = df[col].astype(str).apply(lambda x: re.sub(r'[\t\n\r]', '', str(x))).str.strip()
                df[col] = df[col].replace(['nan', 'None', ''], None)
        return df
    
    def apply_mapping(self, df):
        reverse_mapping = {}
        for standard, actual in self.config["column_mapping"].items():
            if actual and actual in df.columns:
                reverse_mapping[actual] = standard
        return df.rename(columns=reverse_mapping)
    
    def start_complete(self):
        if not self.trans_file_var.get():
            messagebox.showwarning("警告", "请先选择交易明细文件！")
            return
        if not self.acc_file_var.get():
            messagebox.showwarning("警告", "请先选择账户信息文件！")
            return
        
        # ====== 询问是否去重 ======
        result = messagebox.askyesno(
            "去重确认",
            "是否按交易流水号去重？\n\n"
            "选择【是】：保留第一条，删除后续重复记录\n"
            "选择【否】：保留所有记录（包括重复）\n\n"
            "建议：常规分析选择【是】"
        )
        self.deduplicate = result
        
        thread = threading.Thread(target=self.run_complete)
        thread.daemon = True
        thread.start()
    
    def run_complete(self):
        try:
            self.complete_in_progress = True
            self.complete_btn.config(state="disabled", text="⏳ 补全中...")
            self.complete_progress.start()
            
            self.log_complete("="*60)
            self.log_complete("开始补全交易明细数据")
            if self.deduplicate:
                self.log_complete("✓ 去重模式：已启用（按交易流水号去重，保留第一条）")
            else:
                self.log_complete("✓ 去重模式：已关闭（保留所有记录）")
            self.log_complete("="*60)
            
            trans_file = self.trans_file_var.get()
            acc_file = self.acc_file_var.get()
            
            df_trans = pd.read_excel(trans_file)
            df_acc = pd.read_excel(acc_file)
            self.log_complete(f"交易明细: {len(df_trans)}行, 账户信息: {len(df_acc)}行")
            
            df_trans = self.apply_mapping(df_trans)
            df_acc = self.apply_mapping(df_acc)
            
            # 去除无效行
            amount_col = self.config["column_mapping"]["交易金额"]
            if amount_col in df_trans.columns:
                before = len(df_trans)
                df_trans = df_trans[df_trans[amount_col].notna() & (df_trans[amount_col].astype(str).str.strip() != '')]
                self.log_complete(f"删除交易金额为空: {before} -> {len(df_trans)}行")
            
            card_col = self.config["column_mapping"]["交易卡号"]
            account_col = self.config["column_mapping"]["交易账号"]
            if card_col in df_trans.columns and account_col in df_trans.columns:
                before = len(df_trans)
                mask = df_trans[card_col].notna() | df_trans[account_col].notna()
                df_trans = df_trans[mask]
                self.log_complete(f"删除卡号和账号同时为空: {before} -> {len(df_trans)}行")
            
            # 清洗数据
            df_trans = self.clean_dataframe(df_trans)
            df_acc = self.clean_dataframe(df_acc)
            
            # ========== 按交易流水号去重（根据用户选择） ==========
            if self.deduplicate:
                self.log_complete("4. 按交易流水号去重...")
                
                serial_col = self.config["column_mapping"].get("交易流水号", "交易流水号")
                
                if serial_col in df_trans.columns:
                    before_dedup = len(df_trans)
                    duplicate_counts = df_trans[serial_col].value_counts()
                    duplicates = duplicate_counts[duplicate_counts > 1]
                    
                    if len(duplicates) > 0:
                        df_trans = df_trans.drop_duplicates(subset=[serial_col], keep='first')
                        self.log_complete(f"   去重前: {before_dedup} 条，去重后: {len(df_trans)} 条")
                        self.log_complete(f"   删除了 {before_dedup - len(df_trans)} 条重复记录")
                        self.log_complete(f"   发现 {len(duplicates)} 个重复的流水号")
                        self.log_complete(f"   重复最多的流水号出现 {duplicates.max()} 次")
                    else:
                        self.log_complete(f"   ✓ 未发现重复的流水号")
                else:
                    self.log_complete(f"   ⚠️ 未找到交易流水号列，跳过去重")
            else:
                self.log_complete("4. 跳过按交易流水号去重（用户选择保留所有记录）")
            
            # 填充交易卡号
            if card_col in df_trans.columns and account_col in df_trans.columns and account_col in df_acc.columns:
                empty_card = df_trans[card_col].isna()
                card_map = df_acc[df_acc[card_col].notna()].set_index(account_col)[card_col].to_dict()
                df_trans.loc[empty_card, card_col] = df_trans.loc[empty_card, account_col].map(card_map)
                filled = df_trans.loc[empty_card, card_col].notna().sum()
                self.log_complete(f"填充交易卡号: {filled}/{empty_card.sum()}条")
            
            # 填充账户信息
            customer_col = self.config["column_mapping"]["交易方户名"]
            name_col = self.config["column_mapping"]["账户开户名称"]
            id_col = self.config["column_mapping"]["开户人证件号码"]
            
            if customer_col in df_trans.columns:
                empty_customer = df_trans[customer_col].isna()
                if empty_customer.sum() > 0:
                    acc_map = {}
                    for _, row in df_acc.iterrows():
                        acc = str(row[account_col]) if pd.notna(row[account_col]) else ''
                        if acc and acc != 'nan':
                            acc_map[acc] = {name_col: row.get(name_col, ''), id_col: row.get(id_col, '')}
                    
                    for col in [name_col, id_col]:
                        if col not in df_trans.columns:
                            df_trans[col] = None
                    
                    for idx in df_trans[empty_customer].index:
                        acc = str(df_trans.at[idx, account_col]) if pd.notna(df_trans.at[idx, account_col]) else ''
                        if acc and acc in acc_map:
                            df_trans.at[idx, name_col] = acc_map[acc][name_col]
                            df_trans.at[idx, id_col] = acc_map[acc][id_col]
                    
                    filled_count = df_trans.loc[empty_customer, name_col].notna().sum()
                    self.log_complete(f"填充账户信息: {filled_count}/{empty_customer.sum()}条")
            
            # 合并列数据
            if customer_col in df_trans.columns and name_col in df_trans.columns:
                merged_names = []
                for idx, row in df_trans.iterrows():
                    customer = str(row[customer_col]) if pd.notna(row[customer_col]) and str(row[customer_col]) != 'nan' else ''
                    acc_name = str(row[name_col]) if pd.notna(row[name_col]) and str(row[name_col]) != 'nan' else ''
                    if acc_name and acc_name != '':
                        merged_names.append(acc_name)
                    elif customer and customer != '':
                        merged_names.append(customer)
                    else:
                        merged_names.append('')
                df_trans[name_col] = merged_names
            
            # 导出结果
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_dir = os.path.dirname(trans_file)
            output_file = os.path.join(output_dir, f"交易明细_补全结果_{timestamp}.xlsx")
            
            priority = ['交易流水号', '交易卡号', '交易账号', '账户开户名称', '开户人证件号码']
            existing = [col for col in priority if col in df_trans.columns]
            other = [col for col in df_trans.columns if col not in existing]
            df_final = df_trans[existing + other]
            
            df_final.to_excel(output_file, index=False)
            self.log_complete(f"✅ 补全完成！输出: {os.path.basename(output_file)}")
            
            messagebox.showinfo("补全完成", f"交易明细补全完成！\n总行数: {len(df_final)}\n文件已保存")
            
        except Exception as e:
            self.log_complete(f"❌ 补全失败: {e}")
            messagebox.showerror("错误", f"补全失败:\n{e}")
        finally:
            self.complete_progress.stop()
            self.complete_btn.config(state="normal", text="🔧 开始补全数据")
            self.complete_in_progress = False
    # ================== 分析功能 ==================
    def start_analysis(self):
        if not self.current_file:
            messagebox.showwarning("警告", "请先选择要分析的Excel文件！")
            return
        
        thread = threading.Thread(target=self.run_analysis)
        thread.daemon = True
        thread.start()
    
    def run_analysis(self):
        try:
            self.analyze_btn.config(state="disabled", text="⏳ 分析中...")
            self.progress.start()
            
            night_start = int(self.night_start.get())
            night_end = int(self.night_end.get())
            min_amount = int(self.min_amount.get())
            
            df = pd.read_excel(self.current_file)
            df = self.apply_mapping(df)
            
            time_col = self.config["column_mapping"]["交易时间"]
            amount_col = self.config["column_mapping"]["交易金额"]
            
            df['交易时间_格式化'] = pd.to_datetime(df[time_col], errors='coerce')
            df['小时'] = df['交易时间_格式化'].dt.hour
            
            if night_start > night_end:
                df['夜间交易'] = ((df['小时'] >= night_start) | (df['小时'] <= night_end)).astype(int)
            else:
                df['夜间交易'] = ((df['小时'] >= night_start) & (df['小时'] <= night_end)).astype(int)
            
            night_count = df['夜间交易'].sum()
            
            df['金额_数值'] = pd.to_numeric(df[amount_col], errors='coerce')
            df['整数倍金额交易'] = ((df['金额_数值'] >= min_amount) & (df['金额_数值'] % 100 == 0)).astype(int)
            integer_count = df['整数倍金额交易'].sum()
            
            output_dir = os.path.dirname(self.current_file)
            base_name = os.path.splitext(os.path.basename(self.current_file))[0]
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_folder = os.path.join(output_dir, f"分析结果_{timestamp}")
            os.makedirs(output_folder, exist_ok=True)
            
            group_cols = [c for c in ['交易卡号', '交易账号', '账户开户名称', '开户人证件号码'] if c in df.columns]
            if group_cols:
                night_stats = df[df['夜间交易'] == 1].groupby(group_cols).size().reset_index(name='夜间转账次数')
                night_stats.to_excel(os.path.join(output_folder, f"1_夜间转账统计_{base_name}.xlsx"), index=False)
            
            summary_df = pd.DataFrame({
                '统计项目': ['分析时间', '源文件', '总交易笔数', '夜间交易笔数', '夜间交易占比', 
                           f'{min_amount}元以上整数倍交易笔数', f'{min_amount}元以上整数倍交易占比'],
                '数值': [datetime.now().strftime("%Y-%m-%d %H:%M:%S"), os.path.basename(self.current_file),
                        len(df), night_count, f"{night_count/len(df)*100:.2f}%",
                        integer_count, f"{integer_count/len(df)*100:.2f}%"]
            })
            summary_df.to_excel(os.path.join(output_folder, f"2_统计分析报告_{base_name}.xlsx"), index=False)
            
            self.log_analysis(f"✅ 分析完成！输出: {output_folder}")
            messagebox.showinfo("完成", f"分析完成！\n结果已保存到:\n{output_folder}")
            
        except Exception as e:
            self.log_analysis(f"❌ 分析失败: {e}")
            messagebox.showerror("错误", f"分析失败:\n{e}")
        finally:
            self.progress.stop()
            self.analyze_btn.config(state="normal", text="🚀 开始分析")
    # ================== 行为分析 ==================
    def start_behavior_analysis(self):
        if not self.behavior_file_var.get():
            messagebox.showwarning("警告", "请先选择要分析的文件！")
            return
        thread = threading.Thread(target=self.run_behavior_analysis)
        thread.daemon = True
        thread.start()
    
    def run_behavior_analysis(self):
        try:
            self.behavior_btn.config(state="disabled", text="⏳ 分析中...")
            self.behavior_progress.start()
            self.behavior_result_text.delete(1.0, tk.END)
            
            freq_threshold = int(self.freq_threshold.get())
            amount_threshold = int(self.amount_threshold.get())
            time_window = int(self.time_window.get())
            
            df = pd.read_excel(self.behavior_file_var.get())
            df = self.apply_mapping(df)
            
            time_col = self.config["column_mapping"]["交易时间"]
            amount_col = self.config["column_mapping"]["交易金额"]
            account_col = self.config["column_mapping"]["交易账号"]
            
            df['交易时间_格式化'] = pd.to_datetime(df[time_col], errors='coerce')
            df['金额_数值'] = pd.to_numeric(df[amount_col], errors='coerce')
            
            self.log_behavior("="*60)
            self.log_behavior("账户行为分析报告")
            
            if account_col in df.columns:
                account_groups = df.groupby(account_col)
                self.log_behavior(f"\n共有 {len(account_groups)} 个活跃账户")
                
                high_freq_accounts = []
                high_amount_accounts = []
                
                for account, group in account_groups:
                    tx_count = len(group)
                    total_amount = group['金额_数值'].sum()
                    
                    if tx_count >= freq_threshold:
                        high_freq_accounts.append({'账户': account, '交易次数': tx_count, '总金额': total_amount})
                    if total_amount >= amount_threshold:
                        high_amount_accounts.append({'账户': account, '交易次数': tx_count, '总金额': total_amount})
                
                self.log_behavior(f"\n🔴 高频交易账户 (≥{freq_threshold}次):")
                for acc in sorted(high_freq_accounts, key=lambda x: x['交易次数'], reverse=True)[:20]:
                    self.log_behavior(f"  {acc['账户']}: {acc['交易次数']}次, 总金额{acc['总金额']:,.2f}元")
                
                self.log_behavior(f"\n💰 大额交易账户 (≥{amount_threshold:,}元):")
                for acc in sorted(high_amount_accounts, key=lambda x: x['总金额'], reverse=True)[:20]:
                    self.log_behavior(f"  {acc['账户']}: {acc['交易次数']}次, 总金额{acc['总金额']:,.2f}元")
                
                if time_window > 0:
                    cutoff_date = datetime.now() - timedelta(days=time_window)
                    recent_txs = df[df['交易时间_格式化'] >= cutoff_date]
                    self.log_behavior(f"\n📅 最近{time_window}天活跃账户数: {recent_txs[account_col].nunique()}")
                
                output_dir = os.path.dirname(self.behavior_file_var.get())
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_file = os.path.join(output_dir, f"账户行为分析_{timestamp}.xlsx")
                
                with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                    if high_freq_accounts:
                        pd.DataFrame(high_freq_accounts).to_excel(writer, sheet_name='高频账户', index=False)
                    if high_amount_accounts:
                        pd.DataFrame(high_amount_accounts).to_excel(writer, sheet_name='大额账户', index=False)
                
                self.log_behavior(f"\n✅ 详细报告已保存: {output_file}")
            
            messagebox.showinfo("完成", "账户行为分析完成！")
            
        except Exception as e:
            self.log_behavior(f"❌ 分析失败: {e}")
            messagebox.showerror("错误", f"分析失败:\n{e}")
        finally:
            self.behavior_progress.stop()
            self.behavior_btn.config(state="normal", text="🔍 开始行为分析")
    
    # ================== 关联分析 ==================
    def start_relation_analysis(self):
        if not self.relation_file_var.get():
            messagebox.showwarning("警告", "请先选择要分析的文件！")
            return
        thread = threading.Thread(target=self.run_relation_analysis)
        thread.daemon = True
        thread.start()
    
    def run_relation_analysis(self):
        try:
            self.relation_btn.config(state="disabled", text="⏳ 分析中...")
            self.relation_progress.start()
            self.relation_result_text.delete(1.0, tk.END)
            
            min_count = int(self.min_tx_count.get())
            min_amount = float(self.min_total_amount.get())
            
            df = pd.read_excel(self.relation_file_var.get())
            df = self.apply_mapping(df)
            
            account_col = self.config["column_mapping"]["交易账号"]
            counterparty_col = self.config["column_mapping"]["交易方账号"]
            amount_col = self.config["column_mapping"]["交易金额"]
            
            df['金额_数值'] = pd.to_numeric(df[amount_col], errors='coerce')
            
            self.log_relation("="*60)
            self.log_relation("账户关联分析报告")
            
            relations = defaultdict(lambda: {'count': 0, 'total_amount': 0})
            
            for _, row in df.iterrows():
                from_acc = row[account_col] if pd.notna(row[account_col]) else None
                to_acc = row[counterparty_col] if pd.notna(row[counterparty_col]) else None
                amount = row['金额_数值'] if pd.notna(row['金额_数值']) else 0
                
                if from_acc and to_acc and from_acc != to_acc:
                    key = (from_acc, to_acc)
                    relations[key]['count'] += 1
                    relations[key]['total_amount'] += amount
            
            significant_relations = []
            for (from_acc, to_acc), data in relations.items():
                if data['count'] >= min_count and data['total_amount'] >= min_amount:
                    significant_relations.append({
                        '来源账户': from_acc,
                        '目标账户': to_acc,
                        '交易次数': data['count'],
                        '总金额': data['total_amount']
                    })
            
            significant_relations.sort(key=lambda x: x['交易次数'], reverse=True)
            
            self.log_relation(f"\n🔗 发现 {len(significant_relations)} 个显著关联关系")
            
            for rel in significant_relations[:30]:
                self.log_relation(f"📌 {rel['来源账户']} → {rel['目标账户']}")
                self.log_relation(f"   交易次数: {rel['交易次数']}次, 总金额: {rel['总金额']:,.2f}元")
            
            if significant_relations:
                output_dir = os.path.dirname(self.relation_file_var.get())
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_file = os.path.join(output_dir, f"关联分析结果_{timestamp}.xlsx")
                pd.DataFrame(significant_relations).to_excel(output_file, index=False)
                self.log_relation(f"\n✅ 详细报告已保存: {output_file}")
            
            messagebox.showinfo("完成", f"关联分析完成！发现 {len(significant_relations)} 个关联关系")
            
        except Exception as e:
            self.log_relation(f"❌ 分析失败: {e}")
            messagebox.showerror("错误", f"分析失败:\n{e}")
        finally:
            self.relation_progress.stop()
            self.relation_btn.config(state="normal", text="🔍 开始关联分析")
    
    # ================== 筛选功能 ==================
    def start_filter(self):
        if not self.filter_file_var.get():
            messagebox.showwarning("警告", "请先选择要筛选的文件！")
            return
        thread = threading.Thread(target=self.run_filter)
        thread.daemon = True
        thread.start()
    
    def run_filter(self):
        try:
            self.filter_btn.config(state="disabled", text="⏳ 筛选中...")
            self.filter_progress.start()
            self.filter_result_text.delete(1.0, tk.END)
            
            df = pd.read_excel(self.filter_file_var.get())
            df = self.apply_mapping(df)
            
            original_count = len(df)
            self.log_filter("="*60)
            self.log_filter("高级筛选结果")
            self.log_filter(f"原始数据: {original_count} 条记录")
            
            amount_col = self.config["column_mapping"]["交易金额"]
            df['金额_数值'] = pd.to_numeric(df[amount_col], errors='coerce')
            
            if self.amount_min.get():
                min_amt = float(self.amount_min.get())
                df = df[df['金额_数值'] >= min_amt]
                self.log_filter(f"金额 ≥ {min_amt}: {len(df)} 条")
            
            if self.amount_max.get():
                max_amt = float(self.amount_max.get())
                df = df[df['金额_数值'] <= max_amt]
                self.log_filter(f"金额 ≤ {max_amt}: {len(df)} 条")
            
            time_col = self.config["column_mapping"]["交易时间"]
            if time_col in df.columns:
                df['交易时间_格式化'] = pd.to_datetime(df[time_col], errors='coerce')
                
                if self.date_start.get():
                    start_date = pd.to_datetime(self.date_start.get())
                    df = df[df['交易时间_格式化'] >= start_date]
                    self.log_filter(f"时间 ≥ {self.date_start.get()}: {len(df)} 条")
                
                if self.date_end.get():
                    end_date = pd.to_datetime(self.date_end.get())
                    df = df[df['交易时间_格式化'] <= end_date]
                    self.log_filter(f"时间 ≤ {self.date_end.get()}: {len(df)} 条")
            
            if self.keyword.get():
                keyword = self.keyword.get()
                name_col = self.config["column_mapping"]["交易方户名"]
                summary_col = self.config["column_mapping"]["交易摘要"]
                
                mask = pd.Series([False] * len(df))
                if name_col in df.columns:
                    mask = mask | df[name_col].astype(str).str.contains(keyword, na=False, case=False)
                if summary_col in df.columns:
                    mask = mask | df[summary_col].astype(str).str.contains(keyword, na=False, case=False)
                df = df[mask]
                self.log_filter(f"关键词 '{keyword}': {len(df)} 条")
            
            final_count = len(df)
            self.log_filter(f"\n最终筛选结果: {final_count} 条记录 (保留率: {final_count/original_count*100:.1f}%)")
            
            if final_count > 0:
                self.log_filter("\n前10条记录预览:")
                display_cols = [c for c in ['交易时间', '交易金额', '交易账号', '交易方户名'] if c in df.columns]
                for i, (_, row) in enumerate(df.head(10).iterrows(), 1):
                    preview = f"{i}. " + " | ".join([f"{col}:{str(row[col])[:20]}" for col in display_cols])
                    self.log_filter(preview)
                
                total_amount = df['金额_数值'].sum()
                avg_amount = df['金额_数值'].mean()
                self.filter_stats_label.config(text=f"总计: {final_count} 条 | 总金额: {total_amount:,.2f}元 | 平均: {avg_amount:,.2f}元")
            
            output_dir = os.path.dirname(self.filter_file_var.get())
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = os.path.join(output_dir, f"筛选结果_{timestamp}.xlsx")
            df.to_excel(output_file, index=False)
            self.log_filter(f"\n✅ 筛选结果已保存: {output_file}")
            
            messagebox.showinfo("完成", f"筛选完成！\n找到 {final_count} 条记录")
            
        except Exception as e:
            self.log_filter(f"❌ 筛选失败: {e}")
            messagebox.showerror("错误", f"筛选失败:\n{e}")
        finally:
            self.filter_progress.stop()
            self.filter_btn.config(state="normal", text="🔍 执行筛选")

def main():
    root = tk.Tk()
    app = BankTransactionAnalyzer(root)
    root.mainloop()

if __name__ == "__main__":
    main()
